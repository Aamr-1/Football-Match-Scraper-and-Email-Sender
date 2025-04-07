import requests
from bs4 import BeautifulSoup
import csv
import pandas as pd 
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from io import BytesIO
from dotenv import load_dotenv
import os
import smtplib
from email.message import EmailMessage
from email.mime.base import MIMEBase #to handle  email attachement 
from email import encoders
import schedule
import time
from datetime import datetime
date = datetime.now().strftime("%Y-%m-%d")
page = requests.get(f"https://www.goal.com/en/results/{date}")
match_infos = []

def main(page):
    byte_src = page.content
    source_code = BeautifulSoup(byte_src, "lxml")
    
    championships = source_code.find_all("div", {"class": "fco-competition-section"})
    
    def get_matches_info(championships):
        if championships:
            for champ in championships:
                championship_title = champ.find("span", class_="fco-competition-section__header-name")
                if championship_title:
                    championship_title = championship_title.text.strip()
                else:
                    print("Couldn't find the title for this championship.")
                
                teams = champ.find_all("div", class_="fco-team-name fco-long-name")
                if len(teams) >= 2:
                    team_a = teams[0].text.strip()
                    team_b = teams[1].text.strip()
                    
                    teams_score = champ.find_all("div", class_="fco-match-score")
                    if len(teams_score) >= 2:
                        team_a_score = teams_score[0].text.strip()
                        team_b_score = teams_score[1].text.strip()
                    else:
                        team_a_score = "No score"
                        team_b_score = "No score"
                        print("Couldn't find scores for this match.")
                    
                    match_infos.append([championship_title, team_a, team_b, f"{team_a_score} - {team_b_score}"])
                else:
                    print("Couldn't find both teams for this match.")
        else:
            print("No championship sections found on the page.")

    get_matches_info(championships)

    if match_infos:
        excel_file = BytesIO()  # Creating an Excel file in memory (RAM) instead of saving it to the hard disk
    
        columns = ['championship_title 🏆 :', 'Team a ⚽ :', 'Team b ⚽','score 🎯']
        df = pd.DataFrame(match_infos, columns=columns)

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        
            wb = writer.book    # Access the workbook to adjust formatting 
            ws = writer.sheets['Sheet1']

        for i, column in enumerate(df.columns):  # Adjust column widths
            column_letter = get_column_letter(i+1)
            ws.column_dimensions[column_letter].width = max(25, len(column) + 5)  # Set custom width based on column content

        for row in ws.iter_rows():  # Enable text wrapping and set alignment
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, 
                                                 vertical='center',
                                                 horizontal='left')

        # to Get the bytes content of the Excel file
        excel_file.seek(0)
        excel_data = excel_file.getvalue()
        return excel_data
    else:
        print("No match data found to save.")
        return None


load_dotenv()

#getting the variables 
email_address=os.getenv("email_address")
passcode=os.getenv("APP_pass")



def send_email(excel_data, date ):
    if excel_data is None:
        print("No Excel data to send.")
        return
    
    msg=EmailMessage()
    msg['subject']='here is the matches list  for today 👋'
    msg['From']=email_address
    msg['To']='staamr21@gmail.com'
    msg.set_content('Hey,this is the list of fotball match for {date}! ')


   
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(excel_data)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename=f"matches_list_for_{date}.xlsx"')
    msg.attach(part)
    
    
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(email_address,passcode)
        smtp.send_message(msg)
        print("Email sent successfully!")


def job():
    page = requests.get(f"https://www.goal.com/en/results/{date}")
    excel_data = main(page)  #----------> to  Capture the returned excel_data from main function
    send_email(excel_data, date)


schedule.every().day.at("21:00").do(job)


while True:
    schedule.run_pending()  #----------->hecks if any scheduled jobs are pending
    time.sleep(60) #------------->checking every minute 

